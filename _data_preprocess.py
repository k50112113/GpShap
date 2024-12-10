import openpyxl
import numpy as np
import sys
import os
from _utils import smart_system_identification, get_non_empty_index
import pickle

def pad_feature(feature: list, 
                max_padding: int):
    feature_padded = feature + [np.nan]*(max_padding - len(feature))
    return feature_padded

def get_data_key(num_list: list, 
                 name_list: list) -> str:
    data_key = ""
    for i in range(len(name_list)):
        for j in range(len(name_list[i])):
            data_key += f"{name_list[i][j]}_{num_list[i][j]:.3f}_"
    return data_key[:-1]

def get_data_key_c_string(data_key: str, c: float) -> str:
    data_key_list = data_key.split('_')
    if c != -1:
        return "-".join(data_key_list[::2]) + "_" + f"{c:.2f}" + "_" + "-".join(data_key_list[1::2])
    else:
        return "-".join(data_key_list[::2]) + "_" + "-".join(data_key_list[1::2])
    
def standarize_sq(c: np.ndarray, 
                  q: np.ndarray, 
                  sq: np.ndarray, 
                  q_interp_range: tuple, 
                  n_interp_points: int, 
                  n_segs: int, 
                  normalize_to_q: float,
                  verbose: int) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:

    from scipy.signal import savgol_filter
    
    q_interp = np.linspace(*q_interp_range, n_interp_points, endpoint = True)
    sq_interp = []
    for i_c in range(len(c)):
        non_empty_index = get_non_empty_index(sq[i_c])
        # sort_q_index = np.argsort(q[i_c][non_empty_index])
        q_ = q[i_c][non_empty_index]
        sq_ = sq[i_c][non_empty_index]

        # sq_ = savgol_filter(sq_, 21, 3)
        # dsq_dq = np.diff(sq_) / np.diff(q_)

        avg_length = 15
        cut_right_index = (np.diff(q_, prepend = 0.0) < 0).nonzero()[0]
        if len(cut_right_index) > 0:
            cut_right_index = cut_right_index[0]
            cut_left_index = (q_ > q_[cut_right_index]).nonzero()[0][0]
            cut_right_index = cut_right_index + 3

            sq_at_cut_right_index = sq_[cut_right_index : cut_right_index + 1]
            sq_at_cut_left_index = np.mean(sq_[cut_left_index-(avg_length-1):cut_left_index + 1])
            sq_cut_right_prefactor = sq_at_cut_left_index / sq_at_cut_right_index

            q_  = np.concatenate(( q_[ :cut_left_index],  q_[cut_right_index: ]))
            sq_ = np.concatenate((sq_[ :cut_left_index], sq_[cut_right_index: ] * sq_cut_right_prefactor))

        sq_interp.append(np.expand_dims(np.interp(q_interp, q_, sq_), 0))
    
    sq_interp = np.concatenate(sq_interp)
    normalize_to_q_index = np.argmin((q_interp - normalize_to_q)**2)
    if verbose == 1: print(f"Actual noarmlizing q: {q_interp[normalize_to_q_index]}")
    q_interp  = np.tile(q_interp, (len(sq_interp), 1))

    # dq = np.diff(q_interp, axis = 1)/2.0
    # dq = np.concatenate((dq[:, :1], dq[:, :-1] + dq[:, 1:], dq[:, -1:]), axis = 1)
    # prefactor = np.abs(1.0/np.sum(sq_interp * dq, axis = 1))
    # sq_interp = np.transpose((sq_interp.transpose() * prefactor))
    prefactor = sq_interp[:, normalize_to_q_index: normalize_to_q_index + 1]**-1
    sq_interp = sq_interp * prefactor


    seg_size = sq_interp.shape[1] // n_segs
    q_seg_range = []
    sq_seg = []
    for i_seg in range(n_segs):
        sq_seg_ = sq_interp[:, i_seg * seg_size : (i_seg + 1) * seg_size]

        q_seg_range.append([q_interp[0][i_seg * seg_size], q_interp[0][(i_seg + 1) * seg_size - 1]])
        sq_seg.append(sq_seg_.mean(axis = 1))

    q_seg_range = np.array(q_seg_range)
    sq_seg = np.stack(sq_seg).transpose()

    return q_interp, sq_interp, q_seg_range, sq_seg

def fetch_chemical_features(num_list: list, name_list: list, sample_max_num_cations: int, sample_max_num_anions: int, sample_max_num_solvents: int) -> tuple[np.ndarray, list, bool]:

    from _utils import imide_anion_single_side_chain_length, cation_radius_dict, atomic_mass_dict, anion_valence_dict, solvent_density_dict, solvent_bp_dict, solvent_dielectric_dict, solvent_dipole_dict, solvent_viscosity_dict

    cation  = name_list[0]
    anion   = name_list[1]
    solvent = name_list[2]

    cation_num  = num_list[0]
    anion_num   = num_list[1]
    solvent_num = num_list[2]
    
    # imide-base
    anion_valence = []
    anion_size_avg = []
    anion_size_asym = []
    anion_mass = []
    for a_anion in anion:
        anion_valence.append(anion_valence_dict[a_anion])

        if 'FSI' in a_anion or 'BETI' in a_anion or 'BNTI' in a_anion:
            anion_a = a_anion.split('-')
            if len(anion_a) == 2:
                anion_a, anion_b = anion_a[0], anion_a[1]
            else:
                anion_a, anion_b = anion_a[0], anion_a[0]
            anion_size_avg.append(0.5*(imide_anion_single_side_chain_length[anion_a] + imide_anion_single_side_chain_length[anion_b]))
            anion_size_asym.append(abs(imide_anion_single_side_chain_length[anion_a] - imide_anion_single_side_chain_length[anion_b]))
            anion_mass.append(0.5*(atomic_mass_dict[anion_a] + atomic_mass_dict[anion_b]))
        else:
            anion_mass.append(atomic_mass_dict[a_anion])

    # only deals with one type of cation and anion
    cation_valence = int(sum([anion_num[i_anion]*anion_valence[i_anion] for i_anion in range(len(anion))]) / cation_num[0])
    cation_size = [cation_radius_dict[v][cation_valence] for v in cation]
    cation_valence = [cation_valence]

    cation_anion_fraction = cation_num + anion_num
    cation_anion_fraction_sum = sum(cation_anion_fraction)
    cation_anion_fraction = [float(v) / cation_anion_fraction_sum for v in cation_anion_fraction]
    solvent_num_fraction  = solvent_num
    solvent_num_fraction_sum = sum(solvent_num_fraction)
    solvent_num_fraction = [float(v) / solvent_num_fraction_sum for v in solvent_num_fraction]

    solvent_mass = []
    solvent_density = []
    solvent_bp = []
    solvent_dielectric = []
    solvent_dipole = []
    solvent_viscosity = []
    for a_solvent in solvent:
        solvent_mass.append(atomic_mass_dict[a_solvent])
        solvent_density.append(solvent_density_dict[a_solvent])
        solvent_bp.append(solvent_bp_dict[a_solvent])
        solvent_dielectric.append(solvent_dielectric_dict[a_solvent])
        solvent_dipole.append(solvent_dipole_dict[a_solvent])
        solvent_viscosity.append(solvent_viscosity_dict[a_solvent])

    cation_valence = pad_feature(cation_valence, sample_max_num_cations)
    cation_size = pad_feature(cation_size, sample_max_num_cations)
    anion_valence = pad_feature(anion_valence, sample_max_num_anions)
    anion_size_avg = pad_feature(anion_size_avg, sample_max_num_anions)
    anion_size_asym = pad_feature(anion_size_asym, sample_max_num_anions)
    anion_mass = pad_feature(anion_mass, sample_max_num_anions)
    solvent_mass = pad_feature(solvent_mass, sample_max_num_solvents)
    solvent_density = pad_feature(solvent_density, sample_max_num_solvents)
    solvent_bp = pad_feature(solvent_bp, sample_max_num_solvents)
    solvent_dielectric = pad_feature(solvent_dielectric, sample_max_num_solvents)
    solvent_dipole = pad_feature(solvent_dipole, sample_max_num_solvents)
    solvent_viscosity = pad_feature(solvent_viscosity, sample_max_num_solvents)

    feature = np.array(cation_valence     +
                       cation_size        +
                       anion_valence      + 
                       anion_mass         +
                    #    anion_size_avg     + 
                    #    anion_size_asym    +
                       solvent_mass       +
                       solvent_density    + 
                       solvent_bp         +
                       solvent_dielectric +
                       solvent_dipole     +
                       solvent_viscosity)

    feature_name = []
    feature_name += ["cation valence"] * len(cation_valence)
    feature_name += ["cation size"] * len(cation_size)
    feature_name += ["anion valence"] * len(anion_valence)
    feature_name += ["anion mass"] * len(anion_mass)
    # feature_name += ["FC chain avg"] * len(anion_size_avg)
    # feature_name += ["FC chain asym"] * len(anion_size_asym)
    feature_name += ["solvent mass"] * len(solvent_mass)
    feature_name += ["solvent density"] * len(solvent_density)
    feature_name += ["solvent boiling point"] * len(solvent_bp)
    feature_name += ["solvent dielectric"] * len(solvent_dielectric)
    feature_name += ["solvent dipole"] * len(solvent_dipole)
    feature_name += ["solvent viscosity"] * len(solvent_viscosity)

    return feature, feature_name

def standarize_q_data(target_data_dict: dict, data: list, data_key_index_dict: dict, data_parm: dict, verbose: int) -> dict:
    processed_data_dict = {}
    q_range = [float(v) for v in data_parm["q_range"]]
    number_of_interpolation_points = data_parm["number_of_interpolation_points"]
    number_of_segments = data_parm["number_of_segments"]
    normalize_to_q = data_parm["normalize_to_q"]

    for data_key in target_data_dict.keys():
        data_key_index = data_key_index_dict[data_key]
        _, _, _, q, _ = data[data_key_index]

        q_range[0] = max(q_range[0], q.min())
        q_range[1] = min(q_range[1], q.max())

    if verbose == 1:
        print(f"\tnumber of interpolation points: {number_of_interpolation_points}")
        print(f"\tnumber of segments: {number_of_segments}")
        user_q_range = [float(v) for v in data_parm["q_range"]]
        print(f"\tuser-defined q_range: ({user_q_range[0]}, {user_q_range[1]})")
        print(f"\tdata q_range: ({q_range[0]}, {q_range[1]})")

    for i_data_key, data_key in enumerate(target_data_dict.keys()):
        target_c = target_data_dict[data_key]
        data_key_index = data_key_index_dict[data_key]
        _, _, c, q, sq = data[data_key_index]
        c_fg = np.isin(c, target_c)

        assert len(target_c) == len(c_fg.nonzero()[0]), f"Concentration missing, target: {target_c}, avail: {c[c_fg]}"

        q_processed, sq_processed, q_seg_range, sq_seg = standarize_sq(c[c_fg], q[c_fg], sq[c_fg], q_range, number_of_interpolation_points, number_of_segments, normalize_to_q, verbose if i_data_key == 0 else 0)
        processed_data_dict[data_key] = [c[c_fg], q_processed, sq_processed, q_seg_range, sq_seg]
    
    return processed_data_dict

def fetch_chemical_features_data(target_data_dict: dict, data: list, data_key_index_dict: dict, data_parm: dict, verbose: int) -> dict:

    processed_data_dict = {}
    sample_max_num_cations  = 0
    sample_max_num_anions   = 0
    sample_max_num_solvents = 0
    for data_key in target_data_dict.keys():
        data_key_index = data_key_index_dict[data_key]
        d = data[data_key_index]
        if len(d) == 4:
            num_list, name_list, c, p = d
        elif len(d) == 5:
            num_list, name_list, c, q, sq = d
            p = None
        sample_max_num_cations  = max(sample_max_num_cations , len(name_list[0]))
        sample_max_num_anions   = max(sample_max_num_anions  , len(name_list[1]))
        sample_max_num_solvents = max(sample_max_num_solvents, len(name_list[2]))

    if verbose == 1:
        print(f"Fetching chemical features")
        if "target_property" in data_parm:
            target_property = data_parm["target_property"]
            print(f"\ttarget property: {target_property}")

    for data_key in target_data_dict.keys():
        target_c = target_data_dict[data_key]
        data_key_index = data_key_index_dict[data_key]

        d = data[data_key_index]
        if len(d) == 4:
            num_list, name_list, c, p = d
        elif len(d) == 5:
            num_list, name_list, c, q, sq = d
            p = None

        c_fg = np.isin(c, target_c)

        assert len(target_c) == len(c_fg.nonzero()[0]), f"Concentration missing, target: {target_c}, avail: {c[c_fg]}"
        
        c_fg = np.isin(c, target_c)
        feature, feature_name = fetch_chemical_features(num_list, name_list, sample_max_num_cations, sample_max_num_anions, sample_max_num_solvents)

        processed_data_dict[data_key] = [c[c_fg], p[c_fg] if p else None, feature, feature_name]
       
    return processed_data_dict

def preprocess_data(data_type_dict: dict, target_data_dict: dict, data: list, data_key_index_dict: dict, fetch_property: bool, verbose: int) -> tuple[dict, list[str], np.ndarray, np.ndarray, dict, int, int, np.ndarray, np.ndarray, float, float]:

    processed_data_dict = {}

    feature_name_list = []
    X_data = []
    Y_data = []
    XY_data_dict = [] # X_data/Y_data index -> (data_key, c)
    num_samples  = 0.
    num_features = 0.
    X_mean = []
    X_std  = []
    Y_mean = 0.
    Y_std  = 0.

    for data_type in data_type_dict.keys():
        
        if data_type == "property": continue
        data_parm = data_type_dict[data_type]

        if verbose == 1:
            print(f"Preprocessing {data_type.upper()} data")

        processed_data_dict[data_type] = standarize_q_data(target_data_dict, data[data_type], data_key_index_dict[data_type], data_parm, verbose)

    if fetch_property == True:
        data_type = "property"
        data_parm = data_type_dict[data_type]
        processed_data_dict[data_type] = fetch_chemical_features_data(target_data_dict, data[data_type], data_key_index_dict[data_type], data_parm, verbose)
    else:
        for data_type in data_type_dict.keys():

            if data_type == "property": continue
            data_parm = data_type_dict[data_type]

            processed_data_dict["property"] = fetch_chemical_features_data(target_data_dict, data[data_type], data_key_index_dict[data_type], data_parm, verbose)

            break

    for data_key in target_data_dict.keys():
        target_c = target_data_dict[data_key]
        
        data_type = "property"
        c, p, feature, feature_name = processed_data_dict[data_type][data_key]

        c_idx = np.isin(c, target_c).nonzero()[0]

        assert len(target_c) == len(c_idx), f"Concentration missing, target: {target_c}, avail: {c[c_idx]}"

        for i_c in range(len(c_idx)):
            X_data.append(np.expand_dims(np.concatenate((c[c_idx[i_c] : c_idx[i_c] + 1], np.copy(feature))), 0))

            if fetch_property == True:
                assert p, "Property missing"
                Y_data.append(np.expand_dims(p[c_idx[i_c] : c_idx[i_c] + 1], 0))
            else:
                Y_data = None

            XY_data_dict.append((data_key, c[c_idx[i_c]]))

        for data_type in data_type_dict.keys():

            if data_type == "property": continue

            c, q, sq, q_seg_range, sq_seg = processed_data_dict[data_type][data_key]
            c_idx = np.isin(c, target_c).nonzero()[0]

            assert len(target_c) == len(c_idx), f"Concentration missing, target: {target_c}, avail: {c[c_idx]}"

            for i_c in range(len(c_idx)):
                X_data[-len(c_idx) + i_c] = np.concatenate((X_data[-len(c_idx) + i_c], sq_seg[c_idx[i_c] : c_idx[i_c] + 1]), axis = -1)

    feature_name_list = ["concentration"] + feature_name
    for data_type in data_type_dict.keys():
        if data_type == "property": continue
        feature_name_list += [f"{data_type} ({v[0]:.2f} ~ {v[1]:.2f})" for v in  q_seg_range]

    X_data = np.concatenate(X_data)
    num_samples = len(X_data)
    num_features = X_data.shape[1]

    if Y_data:
        Y_data = np.concatenate(Y_data) 
        Y_mean = Y_data.mean()
        Y_std = Y_data.std(ddof = 1)

    for i_feature in range(num_features):
        xtmp = X_data[:, i_feature]
        xtmp = xtmp[np.logical_not(np.isnan(xtmp))]
        X_mean += [xtmp.mean()]
        X_std += [xtmp.std(ddof = 1)]
    X_mean = np.array(X_mean)
    X_std = np.array(X_std)

    # i_feature = 0
    # j_feature = 0
    # while j_feature < num_features:
    #     if feature_name_list[i_feature] == feature_name_list[j_feature]:
    #         j_feature += 1
    #         continue
    #     else:
    #         xtmp = X_data[:, i_feature : j_feature]
    #         xtmp = xtmp[np.logical_not(np.isnan(xtmp))]
    #         X_mean += [xtmp.mean()] * (j_feature - i_feature)
    #         X_std += [xtmp.std(ddof = 1)] * (j_feature - i_feature)
    #         i_feature = j_feature

    # xtmp = X_data[:, i_feature : j_feature]
    # xtmp = xtmp[np.logical_not(np.isnan(xtmp))]
    # X_mean += [xtmp.mean()] * (j_feature - i_feature)
    # X_std += [xtmp.std(ddof = 1)] * (j_feature - i_feature)
    # X_mean = np.array(X_mean)
    # X_std = np.array(X_std)

    duplicate_feature_index = 0
    for i_feature in range(num_features - 1):
        if feature_name_list[i_feature] == feature_name_list[i_feature + 1]:
            duplicate_feature_index += 1
            feature_name_list[i_feature] += f" {duplicate_feature_index}"
        elif duplicate_feature_index > 0:
            duplicate_feature_index += 1
            feature_name_list[i_feature] += f" {duplicate_feature_index}"
            duplicate_feature_index = 0
    if duplicate_feature_index > 0:
        duplicate_feature_index += 1
        feature_name_list[-1] += f" {duplicate_feature_index}"

    for i_feature in range(num_features):
        nan_index = np.isnan(X_data[:, i_feature])
        if len(nan_index.nonzero()[0]) > 0:
            X_data[nan_index, i_feature] = X_mean[i_feature]
    
    zero_index = (X_std == 0).nonzero()[0]
    if len(zero_index) > 0:
        print("warning: these features have zero std: ", *zero_index)
        print("they are set to unity")
        X_std[zero_index] = 1.0

    return processed_data_dict, feature_name_list, X_data, Y_data, XY_data_dict, num_samples, num_features, X_mean, X_std, Y_mean, Y_std 


def load_q_data(filename: str, session_folder: str, data_type: str, only_imide_salt: bool, overwrite: bool = False, verbose: int = 0):
    if os.path.isfile(f"{session_folder}/raw_{data_type}.pickle") and overwrite is False:
        print(f"Pre-load {data_type.upper()} data exists, skip loading raw data.")
        with open(f"{session_folder}/raw_{data_type}.pickle", "rb") as f:
            data, data_key_index_dict = pickle.load(f)
    else:
        data_key_index_dict = {}
        print(f"Loading {data_type.upper()} data from {filename}...")
        data = []
        wb = openpyxl.load_workbook(filename)
        for sheet_name in wb.sheetnames:
            q = []
            sq = []
            # if sheet_name == 'Zn(TFSI)2-H2O-ACN': continue
            if sheet_name == "Table of Content": continue
            sheet = wb[sheet_name]
            for i_row, row in enumerate(sheet.values):
                if i_row == 0:
                    i_col = 0
                    while row[i_col] is None:
                        i_col += 1
                    num_list, name_list, allow = smart_system_identification(row[i_col:i_col+3], only_imide_salt)
                    if allow is False: break
                    if verbose == 1:
                        for i in range(len(name_list)):
                            for j in range(len(name_list[i])):
                                print(f"{name_list[i][j]:10s}:\t{num_list[i][j]}")

                    data_key = get_data_key(num_list, name_list)
                    if data_key not in data_key_index_dict:
                        data.append([num_list, name_list, [], [], []])
                        data_key_index_dict[data_key] = len(data) - 1
                        now_data_index = len(data) - 1
                    else:
                        print(f"warning: Found data with identical chemicals ({data_key}), overwrite with the latter one")
                        data[data_key_index_dict[data_key]] = [num_list, name_list, [], [], []]
                        now_data_index = data_key_index_dict[data_key]
                        
                elif 0 < i_row <= 3:
                    i_col = 0
                    while row[i_col] is None or len(row[i_col].strip()) == 0:
                        i_col += 1
                    if len(q) == 0 and len(row[i_col].strip()) > 1 and row[i_col].strip()[-1] == "m":
                        c = []
                        for v in row:
                            if v is not None and v.strip() != "":
                                c.append(float(v[:v.index('m')]))
                        q  = [[] for _ in range(len(c))]
                        sq = [[] for _ in range(len(c))]
                        continue

                if i_row >= 3 and len(q) > 0:
                    for i in range(2*len(c)):
                        v = row[i] if i < len(row) else None
                        if i % 2 == 0:
                            q[i//2].append(float(v) if v is not None and v != '--' else np.nan)
                        elif i % 2 == 1:
                            sq[i//2].append(float(v) if v is not None and v != '--' else np.nan)
            
            if allow is False:
                continue

            c = np.array(c)
            q = np.array(q)
            sq = np.array(sq)

            if verbose == 1:
                print(f"c\t{c.shape}:\t{c}")
                print(f"q\t{q.shape}")
                print(f"sq\t{sq.shape}")
                print()

            data[now_data_index][2] = c.copy()
            data[now_data_index][3] = q.copy()
            data[now_data_index][4] = sq.copy()

        with open(f"{session_folder}/raw_{data_type}.pickle", "wb") as f:
            pickle.dump((data, data_key_index_dict), f)

    return data, data_key_index_dict

def load_properties(filename: str, session_folder: str, target_property: str, scaling_factor: float, only_imide_salt: bool, overwrite: bool = False, verbose: int = 0):
    if os.path.isfile(f"{session_folder}/raw_property.pickle") and overwrite is False:
        print(f"Pre-load property data exists, skip loading raw data.")
        with open(f"{session_folder}/raw_property.pickle", "rb") as f:
            data, data_key_index_dict, data_col = pickle.load(f)
    else:
        data_key_index_dict = {}
        data = []
        data_col = []
        now_data_index = None
        wb = openpyxl.load_workbook(filename)
        allow = True
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for i_row, row in enumerate(sheet.values):
                if i_row == 0:
                    if len(data_col) > 0:
                        assert all([data_col[i] == row[1+i] for i in range(len(data_col))]), "Data column not consistent"
                    else:
                        i_col = 1
                        while row[i_col] is not None:
                            i_col += 1
                        data_col = row[1:i_col]
                if i_row >= 2:
                    if isinstance(row[0], str):
                        num_list, name_list, allow = smart_system_identification(row[:3], only_imide_salt)
                        if allow is False: continue
                        if verbose == 1:
                            print()
                            for i in range(len(name_list)):
                                for j in range(len(name_list[i])):
                                    print(f"{name_list[i][j]:10s}:\t{num_list[i][j]}")

                        data_key = get_data_key(num_list, name_list)
                        if data_key not in data_key_index_dict:
                            data.append([num_list, name_list, [], []])
                            data_key_index_dict[data_key] = len(data) - 1
                            now_data_index = len(data) - 1
                        else:
                            print(f"warning: Found data with identical chemicals ({data_key}), overwrite with the latter one")
                            data[data_key_index_dict[data_key]] = [num_list, name_list, [], []]
                            now_data_index = data_key_index_dict[data_key]
                    elif allow is True:
                        cdcv = [float(v) if v is not None else float('nan') for v in row[:4]]
                        data[now_data_index][2].append(cdcv[0])
                        data[now_data_index][3].append(cdcv[1:])
                        if verbose == 1:
                            for i in cdcv:
                                print(f"{i:.2e}\t", end = "")
                            print()


        for i in range(len(data)):
            data[i][-2] = np.array(data[i][-2])
            data[i][-1] = np.array(data[i][-1]) * scaling_factor

        with open(f"{session_folder}/raw_property.pickle", "wb") as f:
            pickle.dump((data, data_key_index_dict, data_col), f)
            
    target_property_index = data_col.index(target_property)
    for i in range(len(data)):
        data[i][-1] = data[i][-1][:, target_property_index]
        non_nan_index = np.logical_not(np.isnan(data[i][-1]))
        data[i][-2] = data[i][-2][non_nan_index]
        data[i][-1] = data[i][-1][non_nan_index]

    return data, data_key_index_dict
